from django.shortcuts import render, redirect, get_object_or_404 
"""
This module contains view functions for the home application of the Blossom School project.
Functions:
    - index(request): Renders the index (home) page of the application.
    - login(request): Handles user authentication and logs the user in.
    - logout(request): Logs the user out and redirects to the appropriate page.
    - authenticate(request): Authenticates user credentials for login.
Imports:
    - Django shortcuts for rendering templates, redirecting, and retrieving objects.
    - Django messages framework for displaying notifications.
    - Django authentication decorators and functions for managing user sessions.
    - Local models and forms for handling application-specific data and user input.
"""
from django.contrib import messages 
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate

from .models import *
from .forms import *
from .decorators import unauthenticated_user

# authentications and dashboards 

# views.py - Add these views to your Django app

from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import Sum, Count, Q, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from students.models import Student
from utils.models import Teacher, Attendance, MonthlySalary
from payments.models import (
    Payment, PaymentInstallment, StudentFeeAssignment, 
    FeeStructure, StudentLedger
)
from home.models import ClassRooms, FeeCategory
from Finance.models import Income, Expense


def dashboard_view(request):
    """Main dashboard view"""
    return render(request, 'dashboard.html')


def dashboard_data_api(request):
    """API endpoint to get all dashboard data"""
    
    # Get current date info
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year
    
    # Calculate date ranges
    first_day_of_month = today.replace(day=1)
    last_month = (first_day_of_month - timedelta(days=1)).month
    last_month_year = (first_day_of_month - timedelta(days=1)).year
    
    # STUDENT STATISTICS
    total_students = Student.objects.all().count()
    students_active = Student.objects.filter(is_active=True).count()
    students_by_status = Student.objects.values('status').annotate(count=Count('id'))
    
    # Students enrolled this month
    students_this_month = Student.objects.filter(
        created_at__month=current_month,
        created_at__year=current_year,
        is_active=True
    ).count()
    
    # Students by class
    students_by_class = ClassRooms.objects.annotate(
        student_count=Count('student')
    ).values('class_name', 'student_count')
    
    # TEACHER STATISTICS
    total_teachers = Teacher.objects.all().count()
    active_teachers = Teacher.objects.filter(is_active=True).count()
    
    # New teachers this month
    new_teachers = Teacher.objects.filter(
        created_at__month=current_month,
        created_at__year=current_year,
        is_active=True
    ).count()
    
    # Teacher attendance today
    today_attendance = Attendance.objects.filter(date=today).aggregate(
        present=Count('id', filter=Q(status='present')),
        absent=Count('id', filter=Q(status='absent')),
        half_day=Count('id', filter=Q(status='half_day'))
    )
    
    # PAYMENT STATISTICS
    # Payments this month
    payments_this_month = Payment.objects.filter(
        payment_date__month=current_month,
        payment_date__year=current_year,
        payment_status='completed'
    )
    
    total_revenue = payments_this_month.aggregate(
        total=Sum('net_amount')
    )['total'] or Decimal('0')
    
    # Last month revenue for comparison
    last_month_revenue = Payment.objects.filter(
        payment_date__month=last_month,
        payment_date__year=last_month_year,
        payment_status='completed'
    ).aggregate(total=Sum('net_amount'))['total'] or Decimal('0')
    
    # Calculate percentage change
    if last_month_revenue > 0:
        revenue_change = ((total_revenue - last_month_revenue) / last_month_revenue) * 100
    else:
        revenue_change = 0
    
    # Pending installments
    pending_installments = PaymentInstallment.objects.filter(
        status__in=['pending', 'overdue']
    ).count()
    
    overdue_installments = PaymentInstallment.objects.filter(
        status='overdue'
    ).count()
    
    # Total pending amount
    pending_amount = PaymentInstallment.objects.filter(
        status__in=['pending', 'overdue']
    ).aggregate(
        total=Sum('amount') - Sum('paid_amount')
    )
    
    # Recent payments (last 10)
    recent_payments = Payment.objects.filter(
        payment_status='completed'
    ).select_related('student').order_by('-payment_date')[:10]
    
    recent_payments_list = [{
        'student_name': p.student.get_full_name(),
        'amount': float(p.net_amount),
        'payment_date': p.payment_date.strftime('%Y-%m-%d'),
        'payment_method': p.get_payment_method_display(),
        'payment_id': p.payment_id
    } for p in recent_payments]
    
    # FEE COLLECTION SUMMARY (This Month)
    # Get all installments due this month
    installments_this_month = PaymentInstallment.objects.filter(
        due_date__month=current_month,
        due_date__year=current_year
    )
    
    collected_this_month = installments_this_month.filter(
        status='paid'
    ).aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    
    pending_this_month = installments_this_month.filter(
        status__in=['pending', 'overdue']
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    total_target = collected_this_month + pending_this_month
    collection_percentage = (collected_this_month / total_target * 100) if total_target > 0 else 0
    
    # NEXT MONTH FORECAST (MANDATORY)
    next_month = (today.replace(day=1) + timedelta(days=32)).replace(day=1)
    next_month_end = (next_month.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # Get all installments due next month
    next_month_installments = PaymentInstallment.objects.filter(
        due_date__gte=next_month,
        due_date__lte=next_month_end,
        status__in=['pending', 'partially_paid']
    )
    
    forecast_amount = next_month_installments.aggregate(
        total=Sum('amount') - Sum('paid_amount')
    )
    forecast_total = forecast_amount['total'] or Decimal('0')
    forecast_count = next_month_installments.count()
    
    # Forecast by week
    forecast_by_week = []
    for week in range(1, 5):
        week_start = next_month + timedelta(weeks=week-1)
        week_end = week_start + timedelta(days=6)
        
        week_installments = next_month_installments.filter(
            due_date__gte=week_start,
            due_date__lte=week_end
        )
        
        week_total = week_installments.aggregate(
            total=Sum('amount') - Sum('paid_amount')
        )['total'] or Decimal('0')
        
        forecast_by_week.append({
            'week': f'Week {week}',
            'amount': float(week_total)
        })
    
    # REVENUE & EXPENSE TREND (Last 6 months)
    revenue_trend = []
    expense_trend = []
    months_labels = []
    
    for i in range(5, -1, -1):
        month_date = (today.replace(day=1) - timedelta(days=i*30))
        month = month_date.month
        year = month_date.year
        
        # Revenue from payments
        month_revenue = Payment.objects.filter(
            payment_date__month=month,
            payment_date__year=year,
            payment_status='completed'
        ).aggregate(total=Sum('net_amount'))['total'] or Decimal('0')
        
        # Expenses
        month_expense = Expense.objects.filter(
            date__month=month,
            date__year=year
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        revenue_trend.append(float(month_revenue))
        expense_trend.append(float(month_expense))
        months_labels.append(month_date.strftime('%b'))
    
    # Income from Income model
    income_trend = []
    for i in range(5, -1, -1):
        month_date = (today.replace(day=1) - timedelta(days=i*30))
        month = month_date.month
        year = month_date.year
        
        month_income = Income.objects.filter(
            date__month=month,
            date__year=year
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        income_trend.append(float(month_income))
    
    # EXPENSES THIS MONTH
    expenses_this_month = Expense.objects.filter(
        date__month=current_month,
        date__year=current_year
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Compile all data
    dashboard_data = {
        'students': {
            'total': total_students,
            'active': students_active,
            'this_month': students_this_month,
            'by_status': list(students_by_status),
            'by_class': list(students_by_class)
        },
        'teachers': {
            'total': total_teachers,
            'active': active_teachers,
            'new_this_month': new_teachers,
            'attendance_today': {
                'present': today_attendance['present'] or 0,
                'absent': today_attendance['absent'] or 0,
                'half_day': today_attendance['half_day'] or 0
            }
        },
        'payments': {
            'total_revenue': float(total_revenue),
            'revenue_change': float(revenue_change),
            'pending_installments': pending_installments,
            'overdue_installments': overdue_installments,
            'recent_payments': recent_payments_list,
            'total_payments_count': payments_this_month.count()
        },
        'fee_collection': {
            'collected': float(collected_this_month),
            'pending': float(pending_this_month),
            'total_target': float(total_target),
            'percentage': float(collection_percentage)
        },
        'forecast': {
            'next_month_total': float(forecast_total),
            'installment_count': forecast_count,
            'by_week': forecast_by_week,
            'month_name': next_month.strftime('%B %Y')
        },
        'trends': {
            'labels': months_labels,
            'revenue': revenue_trend,
            'expenses': expense_trend,
            'income': income_trend
        },
        'expenses': {
            'this_month': float(expenses_this_month)
        }
    }
    
    return JsonResponse(dashboard_data)


def get_class_distribution(request):
    """Get class-wise student distribution"""
    classes = ClassRooms.objects.annotate(
        student_count=Count('student', filter=Q(student__is_active=True))
    ).values('class_name', 'student_count')
    
    class_data = []
    for cls in classes:
        # Assuming capacity of 50 per class (adjust as needed)
        capacity = 50
        percentage = (cls['student_count'] / capacity * 100) if capacity > 0 else 0
        
        class_data.append({
            'name': cls['class_name'],
            'students': cls['student_count'],
            'capacity': capacity,
            'percentage': round(percentage, 1)
        })
    
    return JsonResponse({'classes': class_data})


def get_payment_status_chart(request):
    """Get payment status for pie chart"""
    statuses = PaymentInstallment.objects.values('status').annotate(
        count=Count('id')
    )
    
    status_data = {
        'labels': [],
        'data': [],
        'colors': []
    }
    
    color_map = {
        'paid': '#43A574',
        'pending': '#FFB804',
        'overdue': '#FF9587',
        'partially_paid': '#89C0FF'
    }
    
    for status in statuses:
        status_data['labels'].append(status['status'].replace('_', ' ').title())
        status_data['data'].append(status['count'])
        status_data['colors'].append(color_map.get(status['status'], '#214888'))
    
    return JsonResponse(status_data)

from django.db.models import Sum
from django.utils import timezone

@unauthenticated_user
def index(request):
    today = timezone.now()
    students_count = Student.objects.filter(status = "enrolled").count()
    staff_count = Teacher.objects.filter(status = 'active').count()
    total_revenue =  (
        Income.objects
        .filter(date__year=today.year, date__month=today.month)
        .aggregate(total=Sum('amount'))['total'] or 0
    )
    pending_installments = PaymentInstallment.objects.filter(
    due_date__year=today.year,
    due_date__month=today.month
    ).exclude(
        status='paid'
    ).count()

    context = {"students_count":students_count,"staff_count":staff_count,"pending_installments":pending_installments,"total_revenue":total_revenue }
    return render(request,"auth_templates/index.html")


def signin(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username = username, password = password)
        if user is not None:
            login(request,user)
            return redirect("index")
        else:
            messages.error(request, "username or password incorrect")
            return redirect('signin')
    return render(request,"auth_templates/login.html")

def signout(request):
    logout(request)
    return redirect("signin")

# views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import update_session_auth_hash
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
import base64
import os
from PIL import Image
import io
from .forms import ProfileUpdateForm, CustomPasswordChangeForm
from .decorators import unauthenticated_user

@unauthenticated_user
def profile(request):
    """Display user profile page"""
    context = {
        'user': request.user,
        'profile_form': ProfileUpdateForm(instance=request.user),
        'password_form': CustomPasswordChangeForm(user=request.user),
    }
    return render(request, 'auth_templates/profile.html', context)

@unauthenticated_user
def update_profile(request):
    """Handle profile information updates"""
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, instance=request.user)
        
        if form.is_valid():
            form.save()
            # messages.success(request, 'Profile updated successfully!')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Profile updated successfully!'
                })
            
            return redirect('profile')
        else:
            # Handle form errors
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            
            error_message = '; '.join(error_messages)
            messages.error(request, f'Error updating profile: {error_message}')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors,
                    'message': error_message
                })
    
    return redirect('profile')

@unauthenticated_user
def change_password(request):
    """Handle password change requests"""
    if request.method == 'POST':
        form = CustomPasswordChangeForm(data=request.POST, user=request.user)
        
        if form.is_valid():
            user = form.save()
            # Update session to prevent logout
            update_session_auth_hash(request, user)
            
            # messages.success(request, 'Password updated successfully!')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Password updated successfully!'
                })
            
            return redirect('profile')
        else:
            # Handle form errors
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            
            error_message = '; '.join(error_messages)
            # messages.error(request, f'Error changing password: {error_message}')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors,
                    'message': error_message
                })
    
    return redirect('profile')

@unauthenticated_user
@csrf_exempt
def upload_avatar(request):
    """Handle avatar image uploads with proper file handling"""
    if request.method == 'POST':
        try:
            # Handle regular file upload
            if 'avatar' in request.FILES:
                avatar_file = request.FILES['avatar']
                
                # Validate file type
                allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif']
                if avatar_file.content_type not in allowed_types:
                    return JsonResponse({
                        'success': False,
                        'message': 'Please upload a valid image file (JPEG, PNG, or GIF)'
                    })
                
                # Validate file size (max 5MB)
                if avatar_file.size > 5 * 1024 * 1024:
                    return JsonResponse({
                        'success': False,
                        'message': 'File size must be less than 5MB'
                    })
                
                try:
                    # Process image with PIL for validation and optimization
                    image = Image.open(avatar_file)
                    
                    # Validate image dimensions
                    if image.width > 2000 or image.height > 2000:
                        return JsonResponse({
                            'success': False,
                            'message': 'Image dimensions must be less than 2000x2000 pixels'
                        })
                    
                    if image.width < 50 or image.height < 50:
                        return JsonResponse({
                            'success': False,
                            'message': 'Image must be at least 50x50 pixels'
                        })
                    
                    # Convert to RGB if necessary (for JPEG)
                    if image.mode in ('RGBA', 'LA'):
                        background = Image.new('RGB', image.size, (255, 255, 255))
                        background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
                        image = background
                    elif image.mode != 'RGB':
                        image = image.convert('RGB')
                    
                    # Resize if too large (maintain aspect ratio)
                    max_size = (800, 800)
                    if image.width > max_size[0] or image.height > max_size[1]:
                        image.thumbnail(max_size, Image.Resampling.LANCZOS)
                    
                    # Save processed image to memory
                    img_buffer = io.BytesIO()
                    image.save(img_buffer, format='JPEG', quality=85, optimize=True)
                    img_buffer.seek(0)
                    
                    # Create a new ContentFile with processed image
                    processed_file = ContentFile(img_buffer.read())
                    
                    # Generate unique filename
                    file_extension = 'jpg'  # Always save as JPEG after processing
                    file_name = f'avatars/user_{request.user.id}_avatar.{file_extension}'
                    
                    # Remove old avatar if exists
                    if hasattr(request.user, 'avatar') and request.user.avatar:
                        try:
                            default_storage.delete(request.user.avatar.name)
                        except:
                            pass  # Ignore if file doesn't exist
                    
                    # Save new file
                    saved_path = default_storage.save(file_name, processed_file)
                    
                    # Update user model (assuming you have an avatar field)
                    # If you don't have an avatar field in your user model, you'll need to add it
                    # or create a separate UserProfile model with avatar field
                    try:
                        request.user.avatar = saved_path
                        request.user.save()
                    except Exception as e:
                        # If avatar field doesn't exist, just return success without saving to model
                        pass
                    
                    return JsonResponse({
                        'success': True,
                        'avatar_url': default_storage.url(saved_path),
                        'message': 'Avatar updated successfully!'
                    })
                    
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'message': f'Error processing image: {str(e)}'
                    })
            
            # Handle base64 image data from JavaScript (if needed for future use)
            elif 'avatar_data' in request.POST:
                try:
                    image_data = request.POST['avatar_data']
                    # Remove data:image/jpeg;base64, prefix
                    format_info, imgstr = image_data.split(';base64,') 
                    ext = format_info.split('/')[-1]
                    
                    # Decode base64 image
                    img_data = base64.b64decode(imgstr)
                    
                    # Process with PIL
                    image = Image.open(io.BytesIO(img_data))
                    
                    # Same processing as above
                    if image.mode in ('RGBA', 'LA'):
                        background = Image.new('RGB', image.size, (255, 255, 255))
                        background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
                        image = background
                    elif image.mode != 'RGB':
                        image = image.convert('RGB')
                    
                    max_size = (800, 800)
                    if image.width > max_size[0] or image.height > max_size[1]:
                        image.thumbnail(max_size, Image.Resampling.LANCZOS)
                    
                    img_buffer = io.BytesIO()
                    image.save(img_buffer, format='JPEG', quality=85, optimize=True)
                    img_buffer.seek(0)
                    
                    processed_file = ContentFile(img_buffer.read())
                    file_name = f'avatars/user_{request.user.id}_avatar.jpg'
                    
                    # Remove old avatar
                    if hasattr(request.user, 'avatar') and request.user.avatar:
                        try:
                            default_storage.delete(request.user.avatar.name)
                        except:
                            pass
                    
                    saved_path = default_storage.save(file_name, processed_file)
                    
                    try:
                        request.user.avatar = saved_path
                        request.user.save()
                    except:
                        pass
                    
                    return JsonResponse({
                        'success': True,
                        'avatar_url': default_storage.url(saved_path),
                        'message': 'Avatar updated successfully!'
                    })
                    
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'message': f'Error processing base64 image: {str(e)}'
                    })
            
            return JsonResponse({
                'success': False,
                'message': 'No image data received'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error uploading avatar: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    })

@unauthenticated_user
def get_user_activity(request):
    """Get user activity data for the profile page"""
    # This is a placeholder - implement based on your activity tracking needs
    activities = [
        {
            'type': 'login',
            'title': 'Last Login',
            'time': request.user.last_login.strftime('%B %d, %Y at %I:%M %p') if request.user.last_login else 'Never',
            'icon': 'fas fa-sign-in-alt'
        },
        {
            'type': 'update',
            'title': 'Profile Created',
            'time': request.user.date_joined.strftime('%B %d, %Y'),
            'icon': 'fas fa-user-edit'
        },
        {
            'type': 'security',
            'title': 'Account Status',
            'time': 'Active' if request.user.is_active else 'Inactive',
            'icon': 'fas fa-shield-alt'
        }
    ]
    
    return JsonResponse({
        'success': True,
        'activities': activities
    })




# settings for school
# 
# 
# 
#  

#site setting and class room Update

@unauthenticated_user 
def site_setting(request):
    class_rooms = ClassRooms.objects.all().order_by('-id')
    fee_category = FeeCategory.objects.all().order_by("-id")
    if request.method == "POST":
        try:
            name = request.POST['class_room']
            class_room = ClassRooms.objects.create(class_name = name, created_by = request.user)
            class_room.save()
            messages.success(request, "Class room saved successfully...")
        except:
            messages.info(request, 'Something wrong....')

        return redirect("site_setting")
    else:
        context = {"class_rooms":class_rooms,"fee_category":fee_category}
        return render(request,"settings/settings.html",context)
    
@unauthenticated_user
def update_class(request, pk):
    class_room = get_object_or_404(ClassRooms, id = pk)
    if request.method == "POST":
        name = request.POST['class_room']
        class_room.class_name = name 
        class_room.save()
        messages.info(request, 'Class Updated.....')
    return redirect("site_setting")

@unauthenticated_user
def delete_class(request, pk):
    class_room = get_object_or_404(ClassRooms, id = pk)
    class_room.delete()
    messages.info(request,"Class room deleted success.....")
    return redirect(site_setting)

#Fee category Section
@unauthenticated_user
def fee_category(request):
    if request.method == "POST":
        try:
            name = request.POST['name']
            fee_cat = FeeCategory.objects.create(name = name)
            fee_cat.save()
            messages.success(request, "Fee Category saved successfully...")
        except:
            messages.info(request, 'Something wrong....')

        return redirect("site_setting")
    return redirect("site_setting")


@unauthenticated_user
def update_fee_category(request, pk):
    cat = get_object_or_404(FeeCategory, id = pk)
    if request.method == "POST":
        name = request.POST['name']
        cat.name = name 
        cat.save()
        messages.info(request, 'Fee Category Updated.....')
    return redirect("site_setting")


@unauthenticated_user
def delete_fee_category(request, pk):
    cat = get_object_or_404(FeeCategory, id = pk)
    cat.delete()
    messages.info(request,"Fee Category deleted success.....")
    return redirect("site_setting")
    







# all report generation 

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
from decimal import Decimal

from students.models import Student, StudentNote
from payments.models import (
    Payment, PaymentItem, StudentLedger, FeeStructure, 
    StudentFeeAssignment, PaymentInstallment
)
from utils.models import Teacher, MonthlySalary, Attendance
from home.models import  FeeCategory
from Finance.models import  Income, Expense




# def generate_fee_tracking_excel(fee_category_id, start_date, end_date):
#     """Generate fee tracking Excel report"""
#     fee_category = FeeCategory.objects.get(id=fee_category_id)
#     start = datetime.strptime(start_date, '%Y-%m-%d').date()
#     end = datetime.strptime(end_date, '%Y-%m-%d').date()
    
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = f"{fee_category.name} Tracking"
    
#     # School Header
#     ws.merge_cells('A1:H1')
#     ws['A1'] = "Blossom British School"
#     ws['A1'].font = Font(size=16, bold=True, color="214888")
#     ws['A1'].alignment = Alignment(horizontal='center')
    
#     ws.merge_cells('A2:H2')
#     ws['A2'] = "Villa No 2 University Street, Ajman UAE"
#     ws['A2'].alignment = Alignment(horizontal='center')
    
#     ws.merge_cells('A3:H3')
#     ws['A3'] = f"Fee Tracking Report: {fee_category.name}"
#     ws['A3'].font = Font(size=14, bold=True)
#     ws['A3'].alignment = Alignment(horizontal='center')
    
#     ws.merge_cells('A4:H4')
#     ws['A4'] = f"Period: {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"
#     ws['A4'].alignment = Alignment(horizontal='center')
    
#     row = 6
    
#     # Headers
#     headers = ['Student ID', 'Student Name', 'Class', 'Total Due', 'Paid Amount', 'Balance', 'Payment Date', 'Status']
#     for col, header in enumerate(headers, 1):
#         cell = ws.cell(row=row, column=col, value=header)
#         cell.font = Font(bold=True, color="FFFFFF")
#         cell.fill = PatternFill(start_color="214888", end_color="214888", fill_type="solid")
#         cell.alignment = Alignment(horizontal='center')
#     row += 1
    
#     # Get all students with this fee category
#     fee_assignments = StudentFeeAssignment.objects.filter(
#         fee_structure__fee_category=fee_category,
#         is_active=True
#     ).select_related('student', 'fee_structure')
    
#     total_due = Decimal('0')
#     total_paid = Decimal('0')
    
#     for assignment in fee_assignments:
#         student = assignment.student
#         due_amount = assignment.get_final_amount()
        
#         # Calculate paid amount for this fee category
#         paid = PaymentItem.objects.filter(
#             payment__student=student,
#             fee_category=fee_category,
#             payment__payment_date__gte=start,
#             payment__payment_date__lte=end,
#             payment__payment_status='completed'
#         ).aggregate(total=Sum('net_amount'))['total'] or Decimal('0')
        
#         balance = due_amount - paid
        
#         # Get last payment date
#         last_payment = Payment.objects.filter(
#             student=student,
#             payment_items__fee_category=fee_category,
#             payment_date__gte=start,
#             payment_date__lte=end
#         ).order_by('-payment_date').first()
        
#         payment_date = last_payment.payment_date.strftime('%d %b %Y') if last_payment else 'N/A'
#         status = 'Paid' if balance <= 0 else 'Pending' if paid == 0 else 'Partial'
        
#         ws.cell(row=row, column=1, value=student.student_id)
#         ws.cell(row=row, column=2, value=student.get_full_name())
#         ws.cell(row=row, column=3, value=student.class_room.class_name if student.class_room else 'N/A')
#         ws.cell(row=row, column=4, value=float(due_amount))
#         ws.cell(row=row, column=5, value=float(paid))
#         ws.cell(row=row, column=6, value=float(balance))
#         ws.cell(row=row, column=7, value=payment_date)
#         ws.cell(row=row, column=8, value=status)
        
#         # Color code status
#         status_cell = ws.cell(row=row, column=8)
#         if status == 'Paid':
#             status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
#         elif status == 'Pending':
#             status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#         else:
#             status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
#         total_due += due_amount
#         total_paid += paid
#         row += 1
    
#     # Totals
#     row += 1
#     ws.cell(row=row, column=3, value="TOTALS:").font = Font(bold=True, size=12)
#     ws.cell(row=row, column=4, value=float(total_due)).font = Font(bold=True, size=12)
#     ws.cell(row=row, column=5, value=float(total_paid)).font = Font(bold=True, size=12)
#     ws.cell(row=row, column=6, value=float(total_due - total_paid)).font = Font(bold=True, size=12)
    
#     # Adjust column widths
#     for col in range(1, 9):
#         ws.column_dimensions[get_column_letter(col)].width = 16
    
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename="Fee_Tracking_{fee_category.name}_{start}_to_{end}.xlsx"'
#     wb.save(response)
#     return response


# def generate_fee_tracking_pdf(fee_category_id, start_date, end_date):
#     """Generate fee tracking PDF report"""
#     fee_category = FeeCategory.objects.get(id=fee_category_id)
#     start = datetime.strptime(start_date, '%Y-%m-%d').date()
#     end = datetime.strptime(end_date, '%Y-%m-%d').date()
    
#     buffer = BytesIO()
#     doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
    
#     elements = []
#     styles = getSampleStyleSheet()
    
#     # Custom styles
#     title_style = ParagraphStyle(
#         'CustomTitle',
#         parent=styles['Heading1'],
#         fontSize=18,
#         textColor=colors.HexColor('#214888'),
#         alignment=TA_CENTER,
#         spaceAfter=10
#     )
    
#     heading_style = ParagraphStyle(
#         'CustomHeading',
#         parent=styles['Heading2'],
#         fontSize=14,
#         textColor=colors.HexColor('#214888'),
#         spaceAfter=10
#     )
    
#     # Header
#     elements.append(Paragraph("Blossom British School", title_style))
#     elements.append(Paragraph("Villa No 2 University Street, Ajman UAE", styles['Normal']))
#     elements.append(Paragraph(f"Fee Tracking Report: {fee_category.name}", heading_style))
#     elements.append(Paragraph(f"Period: {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}", styles['Normal']))
#     elements.append(Spacer(1, 20))
    
#     # Get all students with this fee category
#     fee_assignments = StudentFeeAssignment.objects.filter(
#         fee_structure__fee_category=fee_category,
#         is_active=True
#     ).select_related('student', 'fee_structure')
    
#     # Create table data
#     table_data = [['Student ID', 'Name', 'Class', 'Due', 'Paid', 'Balance', 'Status']]
#     total_due = Decimal('0')
#     total_paid = Decimal('0')
    
#     for assignment in fee_assignments:
#         student = assignment.student
#         due_amount = assignment.get_final_amount()
        
#         # Calculate paid amount
#         paid = PaymentItem.objects.filter(
#             payment__student=student,
#             fee_category=fee_category,
#             payment__payment_date__gte=start,
#             payment__payment_date__lte=end,
#             payment__payment_status='completed'
#         ).aggregate(total=Sum('net_amount'))['total'] or Decimal('0')
        
#         balance = due_amount - paid
#         status = 'Paid' if balance <= 0 else 'Pending' if paid == 0 else 'Partial'
        
#         table_data.append([
#             student.student_id,
#             student.get_full_name()[:25],
#             student.class_room.class_name if student.class_room else 'N/A',
#             f"AED {due_amount:.2f}",
#             f"AED {paid:.2f}",
#             f"AED {balance:.2f}",
#             status
#         ])
        
#         total_due += due_amount
#         total_paid += paid
    
#     # Add totals row
#     table_data.append([
#         '', 'TOTALS', '',
#         f"AED {total_due:.2f}",
#         f"AED {total_paid:.2f}",
#         f"AED {(total_due - total_paid):.2f}",
#         ''
#     ])
    
#     # Create table
#     col_widths = [1.2*inch, 2*inch, 1*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1*inch]
#     table = Table(table_data, colWidths=col_widths)
    
#     # Apply styles
#     table_style = TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#214888')),
#         ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#         ('FONTSIZE', (0, 0), (-1, 0), 10),
#         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#         ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFB804')),
#         ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
#         ('GRID', (0, 0), (-1, -1), 1, colors.black),
#         ('FONTSIZE', (0, 1), (-1, -1), 8)
#     ])
    
#     table.setStyle(table_style)
#     elements.append(table)
    
#     doc.build(elements)
#     buffer.seek(0)
    
#     response = HttpResponse(buffer, content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="Fee_Tracking_{fee_category.name}_{start}_to_{end}.pdf"'
#     return response



def generate_student_report(request):
    """Generate comprehensive student report"""
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        report_format = request.POST.get('format', 'excel')
        
        if report_format == 'excel':
            return generate_student_excel(student_id)
        else:
            return generate_student_pdf(student_id)
    
    return redirect('reports_dashboard')


def generate_student_excel(student_id):
    """Generate comprehensive student report in Excel"""
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        return HttpResponse("Student not found", status=404)
    
    wb = openpyxl.Workbook()
    
    # ===== SHEET 1: STUDENT PROFILE =====
    ws1 = wb.active
    ws1.title = "Student Profile"
    
    # School Header
    ws1.merge_cells('A1:D1')
    ws1['A1'] = "Blossom British School"
    ws1['A1'].font = Font(size=16, bold=True, color="214888")
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    ws1.merge_cells('A2:D2')
    ws1['A2'] = "Villa No 2 University Street, Ajman UAE"
    ws1['A2'].alignment = Alignment(horizontal='center')
    
    ws1.merge_cells('A3:D3')
    ws1['A3'] = "STUDENT PROFILE REPORT"
    ws1['A3'].font = Font(size=14, bold=True)
    ws1['A3'].alignment = Alignment(horizontal='center')
    
    row = 5
    
    # Student Information
    info_fields = [
        ('Student ID:', student.student_id),
        ('Full Name:', student.get_full_name()),
        ('Date of Birth:', student.date_of_birth.strftime('%d %b %Y')),
        ('Gender:', student.get_gender_display()),
        ('Nationality:', student.nationality),
        ('Class:', student.class_room.class_name if student.class_room else 'N/A'),
        ('Status:', student.get_status_display()),
        ('Year of Admission:', student.year_of_admission),
        ('', ''),
    ]
    
    for label, value in info_fields:
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=value)
        row += 1
    
    # Parent Information
    row += 1
    ws1.merge_cells(f'A{row}:D{row}')
    ws1.cell(row=row, column=1, value="PARENT INFORMATION").font = Font(size=12, bold=True, color="FFFFFF")
    ws1.cell(row=row, column=1).fill = PatternFill(start_color="214888", end_color="214888", fill_type="solid")
    row += 1
    
    parent_fields = [
        ('Father Name:', student.father_name),
        ('Father Mobile:', student.father_mobile),
        ('Father Email:', student.father_email),
        ('Mother Name:', student.mother_name),
        ('Mother Mobile:', student.mother_mobile),
        ('Mother Email:', student.mother_email),
        ('Home Address:', student.full_home_address),
        ('City:', student.city),
    ]
    
    for label, value in parent_fields:
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=value)
        row += 1
    
    # Emergency Contacts
    row += 1
    ws1.merge_cells(f'A{row}:D{row}')
    ws1.cell(row=row, column=1, value="EMERGENCY CONTACTS").font = Font(size=12, bold=True, color="FFFFFF")
    ws1.cell(row=row, column=1).fill = PatternFill(start_color="D54395", end_color="D54395", fill_type="solid")
    row += 1
    
    emergency_fields = [
        ('Primary Contact:', student.first_contact_person),
        ('Relationship:', student.first_contact_relationship),
        ('Phone:', student.first_contact_telephone),
        ('', ''),
        ('Secondary Contact:', student.second_contact_person),
        ('Relationship:', student.second_contact_relationship),
        ('Phone:', student.second_contact_telephone),
    ]
    
    for label, value in emergency_fields:
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=value)
        row += 1
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 35
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20
    
    # ===== SHEET 2: FEE STRUCTURE =====
    ws2 = wb.create_sheet("Fee Structure")
    
    # Header
    ws2.merge_cells('A1:G1')
    ws2['A1'] = f"FEE STRUCTURE - {student.get_full_name()}"
    ws2['A1'].font = Font(size=14, bold=True, color="214888")
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    
    # Headers
    headers = ['Fee Category', 'Frequency', 'Base Amount', 'Discount %', 'Discount Amount', 'Final Amount', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="214888", end_color="214888", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Get fee assignments
    fee_assignments = StudentFeeAssignment.objects.filter(
        student=student,
        is_active=True
    ).select_related('fee_structure', 'fee_structure__fee_category')
    
    total_final = Decimal('0')
    
    for assignment in fee_assignments:
        base_amount = assignment.custom_amount or assignment.fee_structure.amount
        final_amount = assignment.get_final_amount()
        
        ws2.cell(row=row, column=1, value=assignment.fee_structure.fee_category.name)
        ws2.cell(row=row, column=2, value=assignment.fee_structure.get_frequency_display())
        ws2.cell(row=row, column=3, value=float(base_amount))
        ws2.cell(row=row, column=4, value=float(assignment.discount_percentage))
        ws2.cell(row=row, column=5, value=float(assignment.discount_amount))
        ws2.cell(row=row, column=6, value=float(final_amount))
        ws2.cell(row=row, column=7, value='Active' if assignment.is_active else 'Inactive')
        
        total_final += final_amount
        row += 1
    
    # Total
    ws2.cell(row=row, column=5, value="TOTAL:").font = Font(bold=True)
    ws2.cell(row=row, column=6, value=float(total_final)).font = Font(bold=True)
    
    # Adjust column widths
    for col in range(1, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    
    # ===== SHEET 3: PAYMENT HISTORY =====
    ws3 = wb.create_sheet("Payment History")
    
    # Header
    ws3.merge_cells('A1:H1')
    ws3['A1'] = f"PAYMENT HISTORY - {student.get_full_name()}"
    ws3['A1'].font = Font(size=14, bold=True, color="214888")
    ws3['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    
    # Headers
    headers = ['Payment ID', 'Date', 'Fee Category', 'Amount', 'Discount', 'Late Fee', 'Net Amount', 'Method']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="43A574", end_color="43A574", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Get payments
    payments = Payment.objects.filter(
        student=student,
        payment_status='completed'
    ).order_by('-payment_date')
    
    total_paid = Decimal('0')
    
    for payment in payments:
        for item in payment.payment_items.all():
            ws3.cell(row=row, column=1, value=payment.payment_id)
            ws3.cell(row=row, column=2, value=payment.payment_date.strftime('%d %b %Y'))
            ws3.cell(row=row, column=3, value=item.fee_category.name)
            ws3.cell(row=row, column=4, value=float(item.amount))
            ws3.cell(row=row, column=5, value=float(item.discount_amount))
            ws3.cell(row=row, column=6, value=float(item.late_fee))
            ws3.cell(row=row, column=7, value=float(item.net_amount))
            ws3.cell(row=row, column=8, value=payment.get_payment_method_display())
            
            total_paid += item.net_amount
            row += 1
    
    # Total
    row += 1
    ws3.cell(row=row, column=6, value="TOTAL PAID:").font = Font(bold=True, size=12)
    ws3.cell(row=row, column=7, value=float(total_paid)).font = Font(bold=True, size=12)
    ws3.cell(row=row, column=7).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Adjust column widths
    for col in range(1, 9):
        ws3.column_dimensions[get_column_letter(col)].width = 15
    
    # ===== SHEET 4: FINANCIAL SUMMARY =====
    ws4 = wb.create_sheet("Financial Summary")
    
    # Header
    ws4.merge_cells('A1:D1')
    ws4['A1'] = f"FINANCIAL SUMMARY - {student.get_full_name()}"
    ws4['A1'].font = Font(size=14, bold=True, color="214888")
    ws4['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    
    # Calculate summary
    total_due = total_final
    outstanding = total_due - total_paid
    
    summary_data = [
        ('Total Fee Assigned:', float(total_due), 'FFB804'),
        ('Total Paid:', float(total_paid), '43A574'),
        ('Outstanding Balance:', float(outstanding), 'D54395' if outstanding > 0 else '43A574'),
        ('Payment Status:', 'Fully Paid' if outstanding <= 0 else f'AED {outstanding:.2f} Pending', None),
    ]
    
    for label, value, color in summary_data:
        ws4.cell(row=row, column=1, value=label).font = Font(bold=True, size=12)
        cell = ws4.cell(row=row, column=2, value=value if isinstance(value, str) else f"AED {value:.2f}")
        cell.font = Font(size=11)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 2
    
    # Notes
    row += 2
    ws4.merge_cells(f'A{row}:D{row}')
    ws4.cell(row=row, column=1, value="NOTES").font = Font(size=12, bold=True, color="FFFFFF")
    ws4.cell(row=row, column=1).fill = PatternFill(start_color="214888", end_color="214888", fill_type="solid")
    row += 1
    
    notes = StudentNote.objects.filter(student=student).order_by('-created_at')[:5]
    for note in notes:
        ws4.cell(row=row, column=1, value=f"{note.created_at.strftime('%d %b %Y')}:")
        ws4.cell(row=row, column=2, value=note.note)
        row += 1
    
    # Adjust column widths
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 40
    
    # Save workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Student_Report_{student.student_id}_{student.get_full_name()}.xlsx"'
    wb.save(response)
    return response


def generate_student_pdf(student_id):
    """Generate comprehensive student report in PDF"""
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        return HttpResponse("Student not found", status=404)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=50, bottomMargin=40)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#214888'),
        alignment=TA_CENTER,
        spaceAfter=10
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#214888'),
        spaceAfter=10,
        spaceBefore=15
    )
    
    # Header
    elements.append(Paragraph("Blossom British School", title_style))
    elements.append(Paragraph("Villa No 2 University Street, Ajman UAE", styles['Normal']))
    elements.append(Paragraph(f"Student Profile Report", heading_style))
    elements.append(Spacer(1, 20))
    
    # Student Information
    elements.append(Paragraph("STUDENT INFORMATION", heading_style))
    
    student_data = [
        ['Student ID:', student.student_id],
        ['Full Name:', student.get_full_name()],
        ['Date of Birth:', student.date_of_birth.strftime('%d %B %Y')],
        ['Gender:', student.get_gender_display()],
        ['Nationality:', student.nationality],
        ['Class:', student.class_room.class_name if student.class_room else 'N/A'],
        ['Status:', student.get_status_display()],
        ['Year of Admission:', str(student.year_of_admission)],
    ]
    
    student_table = Table(student_data, colWidths=[2.5*inch, 3.5*inch])
    student_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    elements.append(student_table)
    elements.append(Spacer(1, 15))
    
    # Parent Information
    elements.append(Paragraph("PARENT INFORMATION", heading_style))
    
    parent_data = [
        ['Father Name:', student.father_name],
        ['Father Mobile:', student.father_mobile],
        ['Father Email:', student.father_email],
        ['Mother Name:', student.mother_name],
        ['Mother Mobile:', student.mother_mobile],
        ['Mother Email:', student.mother_email],
    ]
    
    parent_table = Table(parent_data, colWidths=[2.5*inch, 3.5*inch])
    parent_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(parent_table)
    elements.append(Spacer(1, 15))
    
    # Fee Structure
    elements.append(Paragraph("FEE STRUCTURE", heading_style))
    
    fee_assignments = StudentFeeAssignment.objects.filter(
        student=student,
        is_active=True
    ).select_related('fee_structure', 'fee_structure__fee_category')
    
    fee_data = [['Fee Category', 'Frequency', 'Amount', 'Discount', 'Final Amount']]
    total_final = Decimal('0')
    
    for assignment in fee_assignments:
        final_amount = assignment.get_final_amount()
        fee_data.append([
            assignment.fee_structure.fee_category.name,
            assignment.fee_structure.get_frequency_display(),
            f"AED {(assignment.custom_amount or assignment.fee_structure.amount):.2f}",
            f"{assignment.discount_percentage}%",
            f"AED {final_amount:.2f}"
        ])
        total_final += final_amount
    
    fee_data.append(['', '', '', 'TOTAL:', f"AED {total_final:.2f}"])
    
    fee_table = Table(fee_data, colWidths=[1.8*inch, 1.3*inch, 1.2*inch, 1*inch, 1.2*inch])
    fee_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#214888')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFB804')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9)
    ]))
    
    elements.append(fee_table)
    elements.append(PageBreak())
    
    # Payment History
    elements.append(Paragraph("PAYMENT HISTORY", heading_style))
    
    payments = Payment.objects.filter(
        student=student,
        payment_status='completed'
    ).order_by('-payment_date')[:15]  # Last 15 payments
    
    payment_data = [['Date', 'Payment ID', 'Category', 'Amount', 'Method']]
    total_paid = Decimal('0')
    
    for payment in payments:
        for item in payment.payment_items.all():
            payment_data.append([
                payment.payment_date.strftime('%d-%b-%y'),
                payment.payment_id,
                item.fee_category.name[:15],
                f"AED {item.net_amount:.2f}",
                payment.get_payment_method_display()[:10]
            ])
            total_paid += item.net_amount
    
    payment_data.append(['', '', '', f"AED {total_paid:.2f}", ''])
    
    payment_table = Table(payment_data, colWidths=[1*inch, 1.4*inch, 1.6*inch, 1.2*inch, 1.3*inch])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#43A574')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6E6E6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 8)
    ]))
    
    elements.append(payment_table)
    elements.append(Spacer(1, 20))
    
    # Financial Summary
    elements.append(Paragraph("FINANCIAL SUMMARY", heading_style))
    
    outstanding = total_final - total_paid
    
    summary_data = [
        ['Total Fee Assigned:', f"AED {total_final:.2f}"],
        ['Total Paid:', f"AED {total_paid:.2f}"],
        ['Outstanding Balance:', f"AED {outstanding:.2f}"],
        ['Payment Status:', 'Fully Paid' if outstanding <= 0 else f'AED {outstanding:.2f} Pending'],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.yellow if outstanding > 0 else colors.lightgreen),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    
    elements.append(summary_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Student_Report_{student.student_id}_{student.get_full_name()}.pdf"'



def reports_dashboard(request):
    """Main reports dashboard"""
    from students.models import Student
    context = {
        'fee_categories': FeeCategory.objects.all().order_by('name'),
        'current_year': timezone.now().year,
        'current_date': timezone.now().date(),
        'students': Student.objects.filter(is_active=True).order_by('first_name', 'last_name'),
    }
    return render(request, 'auth_templates/reports-dashboard.html', context)



def generate_daily_report(request):
    """Generate daily financial report"""
    if request.method == 'POST':
        report_date = request.POST.get('report_date')
        report_format = request.POST.get('format', 'excel')
        
        if report_format == 'excel':
            return generate_daily_excel(report_date)
        else:
            return generate_daily_pdf(report_date)
    
    return redirect('reports_dashboard')


def generate_daily_excel(report_date):
    """Generate daily report in Excel format"""
    date_obj = datetime.strptime(report_date, '%Y-%m-%d').date()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Daily Report {date_obj}"
    
    # School Header
    ws.merge_cells('A1:F1')
    school_name = ws['A1']
    school_name.value = "Blossom British School"
    school_name.font = Font(size=16, bold=True, color="214888")
    school_name.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    address = ws['A2']
    address.value = "Villa No 2 University Street, Ajman UAE"
    address.font = Font(size=10)
    address.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:F3')
    report_title = ws['A3']
    report_title.value = f"Daily Financial Report - {date_obj.strftime('%d %B %Y')}"
    report_title.font = Font(size=14, bold=True)
    report_title.alignment = Alignment(horizontal='center')
    
    row = 5
    
    # Fee Payments Section
    ws[f'A{row}'] = "FEE PAYMENTS"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="214888", end_color="214888", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    headers = ['Student ID', 'Student Name', 'Fee Category', 'Amount', 'Payment Method', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    row += 1
    
    payments = Payment.objects.filter(payment_date=date_obj, payment_status='completed')
    total_fees = Decimal('0')
    
    for payment in payments:
        for item in payment.payment_items.all():
            ws.cell(row=row, column=1, value=payment.student.student_id)
            ws.cell(row=row, column=2, value=payment.student.get_full_name())
            ws.cell(row=row, column=3, value=item.fee_category.name)
            ws.cell(row=row, column=4, value=float(item.net_amount))
            ws.cell(row=row, column=5, value=payment.get_payment_method_display())
            ws.cell(row=row, column=6, value=payment.get_payment_status_display())
            total_fees += item.net_amount
            row += 1
    
    ws.cell(row=row, column=3, value="Total Fee Collection:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(total_fees)).font = Font(bold=True)
    row += 2
    
    # Other Income Section
    ws[f'A{row}'] = "OTHER INCOME"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="43A574", end_color="43A574", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    headers = ['Date', 'Particulars', 'Amount', 'Bill Number', 'Other Details']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    row += 1
    
    incomes = Income.objects.filter(date=date_obj)
    total_income = 0
    
    for income in incomes:
        ws.cell(row=row, column=1, value=income.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=income.perticulers)
        ws.cell(row=row, column=3, value=income.amount)
        ws.cell(row=row, column=4, value=income.bill_number)
        ws.cell(row=row, column=5, value=income.other or '')
        total_income += income.amount
        row += 1
    
    ws.cell(row=row, column=2, value="Total Other Income:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_income).font = Font(bold=True)
    row += 2
    
    # Expenses Section
    ws[f'A{row}'] = "EXPENSES"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="D54395", end_color="D54395", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    headers = ['Date', 'Particulars', 'Amount', 'Bill Number', 'Other Details']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    row += 1
    
    expenses = Expense.objects.filter(date=date_obj)
    total_expense = 0
    
    for expense in expenses:
        ws.cell(row=row, column=1, value=expense.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=expense.perticulers)
        ws.cell(row=row, column=3, value=expense.amount)
        ws.cell(row=row, column=4, value=expense.bill_number)
        ws.cell(row=row, column=5, value=expense.other or '')
        total_expense += expense.amount
        row += 1
    
    ws.cell(row=row, column=2, value="Total Expenses:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_expense).font = Font(bold=True)
    row += 2
    
    # Summary
    ws[f'A{row}'] = "DAILY SUMMARY"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FFB804", end_color="FFB804", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    total_receipts = float(total_fees) + total_income
    net_balance = total_receipts - total_expense
    
    ws.cell(row=row, column=2, value="Total Fee Collection:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(total_fees))
    row += 1
    
    ws.cell(row=row, column=2, value="Total Other Income:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_income)
    row += 1
    
    ws.cell(row=row, column=2, value="Total Receipts:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_receipts).font = Font(bold=True)
    row += 1
    
    ws.cell(row=row, column=2, value="Total Expenses:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_expense)
    row += 1
    
    ws.cell(row=row, column=2, value="Net Balance:").font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=net_balance).font = Font(bold=True, size=12)
    ws.cell(row=row, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Daily_Report_{date_obj}.xlsx"'
    wb.save(response)
    return response


def generate_daily_pdf(report_date):
    """Generate daily report in PDF format"""
    date_obj = datetime.strptime(report_date, '%Y-%m-%d').date()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#214888'),
        alignment=TA_CENTER,
        spaceAfter=10
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#214888'),
        spaceAfter=10
    )
    
    # Header
    elements.append(Paragraph("Blossom British School", title_style))
    elements.append(Paragraph("Villa No 2 University Street, Ajman UAE", styles['Normal']))
    elements.append(Paragraph(f"Daily Financial Report - {date_obj.strftime('%d %B %Y')}", heading_style))
    elements.append(Spacer(1, 20))
    
    # Fee Payments
    elements.append(Paragraph("FEE PAYMENTS", heading_style))
    
    payments = Payment.objects.filter(payment_date=date_obj, payment_status='completed')
    payment_data = [['Student ID', 'Student Name', 'Fee Category', 'Amount', 'Method']]
    total_fees = Decimal('0')
    
    for payment in payments:
        for item in payment.payment_items.all():
            payment_data.append([
                payment.student.student_id,
                payment.student.get_full_name(),
                item.fee_category.name,
                f"AED {item.net_amount:.2f}",
                payment.get_payment_method_display()
            ])
            total_fees += item.net_amount
    
    payment_data.append(['', '', 'Total:', f"AED {total_fees:.2f}", ''])
    
    payment_table = Table(payment_data, colWidths=[1.2*inch, 1.8*inch, 1.5*inch, 1.2*inch, 1.2*inch])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#214888')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6E6E6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(payment_table)
    elements.append(Spacer(1, 20))
    
    # Similar sections for Income, Expenses, and Summary...
    # (Code continues with similar structure)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Daily_Report_{date_obj}.pdf"'
    return response



def generate_date_range_report(request):
    """Generate report for date range"""
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        report_format = request.POST.get('format', 'excel')
        
        if report_format == 'excel':
            return generate_range_excel(start_date, end_date)
        else:
            return generate_range_pdf(start_date, end_date)
    
    return redirect('reports_dashboard')


def generate_range_excel(start_date, end_date):
    """Generate date range report in Excel"""
    start = datetime.strptime(start_date, '%Y-%m-%d').date()
    end = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {start} to {end}"
    
    # School Header
    ws.merge_cells('A1:F1')
    school_name = ws['A1']
    school_name.value = "Blossom British School"
    school_name.font = Font(size=16, bold=True, color="214888")
    school_name.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    address = ws['A2']
    address.value = "Villa No 2 University Street, Ajman UAE"
    address.font = Font(size=10)
    address.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:F3')
    report_title = ws['A3']
    report_title.value = f"Financial Report: {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"
    report_title.font = Font(size=14, bold=True)
    report_title.alignment = Alignment(horizontal='center')
    
    row = 5
    
    # Fee Payments Section
    ws[f'A{row}'] = "FEE PAYMENTS"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="214888", end_color="214888", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    headers = ['Date', 'Student Name', 'Fee Category', 'Amount', 'Payment Method', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    row += 1
    
    payments = Payment.objects.filter(
        payment_date__gte=start,
        payment_date__lte=end,
        payment_status='completed'
    ).order_by('payment_date')
    
    total_fees = Decimal('0')
    
    for payment in payments:
        for item in payment.payment_items.all():
            ws.cell(row=row, column=1, value=payment.payment_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=payment.student.get_full_name())
            ws.cell(row=row, column=3, value=item.fee_category.name)
            ws.cell(row=row, column=4, value=float(item.net_amount))
            ws.cell(row=row, column=5, value=payment.get_payment_method_display())
            ws.cell(row=row, column=6, value=payment.get_payment_status_display())
            total_fees += item.net_amount
            row += 1
    
    ws.cell(row=row, column=3, value="Total Fee Collection:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(total_fees)).font = Font(bold=True)
    row += 2
    
    # Other Income Section
    ws[f'A{row}'] = "OTHER INCOME"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="43A574", end_color="43A574", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    headers = ['Date', 'Particulars', 'Amount', 'Bill Number', 'Other']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    row += 1
    
    incomes = Income.objects.filter(date__gte=start, date__lte=end).order_by('date')
    total_income = 0
    
    for income in incomes:
        ws.cell(row=row, column=1, value=income.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=income.perticulers)
        ws.cell(row=row, column=3, value=income.amount)
        ws.cell(row=row, column=4, value=income.bill_number)
        ws.cell(row=row, column=5, value=income.other or '')
        total_income += income.amount
        row += 1
    
    ws.cell(row=row, column=2, value="Total Other Income:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_income).font = Font(bold=True)
    row += 2
    
    # Expenses Section
    ws[f'A{row}'] = "EXPENSES"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="D54395", end_color="D54395", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    headers = ['Date', 'Particulars', 'Amount', 'Bill Number', 'Other']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    row += 1
    
    expenses = Expense.objects.filter(date__gte=start, date__lte=end).order_by('date')
    total_expense = 0
    
    for expense in expenses:
        ws.cell(row=row, column=1, value=expense.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=expense.perticulers)
        ws.cell(row=row, column=3, value=expense.amount)
        ws.cell(row=row, column=4, value=expense.bill_number)
        ws.cell(row=row, column=5, value=expense.other or '')
        total_expense += expense.amount
        row += 1
    
    ws.cell(row=row, column=2, value="Total Expenses:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_expense).font = Font(bold=True)
    row += 2
    
    # Summary
    ws[f'A{row}'] = "PERIOD SUMMARY"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FFB804", end_color="FFB804", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    total_receipts = float(total_fees) + total_income
    net_balance = total_receipts - total_expense
    
    ws.cell(row=row, column=2, value="Total Fee Collection:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(total_fees))
    row += 1
    
    ws.cell(row=row, column=2, value="Total Other Income:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_income)
    row += 1
    
    ws.cell(row=row, column=2, value="Total Receipts:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_receipts).font = Font(bold=True)
    row += 1
    
    ws.cell(row=row, column=2, value="Total Expenses:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_expense)
    row += 1
    
    ws.cell(row=row, column=2, value="Net Balance:").font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=net_balance).font = Font(bold=True, size=12)
    ws.cell(row=row, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Financial_Report_{start}_to_{end}.xlsx"'
    wb.save(response)
    return response


def generate_range_pdf(start_date, end_date):
    """Generate date range report in PDF format"""
    start = datetime.strptime(start_date, '%Y-%m-%d').date()
    end = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#214888'),
        alignment=TA_CENTER,
        spaceAfter=10
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#214888'),
        spaceAfter=10
    )
    
    # Header
    elements.append(Paragraph("Blossom British School", title_style))
    elements.append(Paragraph("Villa No 2 University Street, Ajman UAE", styles['Normal']))
    elements.append(Paragraph(f"Financial Report: {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}", heading_style))
    elements.append(Spacer(1, 20))
    
    # Fee Payments
    elements.append(Paragraph("FEE PAYMENTS", heading_style))
    
    payments = Payment.objects.filter(
        payment_date__gte=start,
        payment_date__lte=end,
        payment_status='completed'
    ).order_by('payment_date')
    
    payment_data = [['Date', 'Student', 'Category', 'Amount', 'Method']]
    total_fees = Decimal('0')
    
    for payment in payments:
        for item in payment.payment_items.all():
            payment_data.append([
                payment.payment_date.strftime('%d-%b'),
                payment.student.get_full_name()[:20],
                item.fee_category.name[:15],
                f"AED {item.net_amount:.2f}",
                payment.get_payment_method_display()[:10]
            ])
            total_fees += item.net_amount
    
    payment_data.append(['', '', 'Total:', f"AED {total_fees:.2f}", ''])
    
    payment_table = Table(payment_data, colWidths=[1.2*inch, 2.5*inch, 1.8*inch, 1.3*inch, 1.3*inch])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#214888')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6E6E6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8)
    ]))
    
    elements.append(payment_table)
    elements.append(Spacer(1, 20))
    
    # Income
    elements.append(Paragraph("OTHER INCOME", heading_style))
    
    incomes = Income.objects.filter(date__gte=start, date__lte=end).order_by('date')
    income_data = [['Date', 'Particulars', 'Amount', 'Bill Number']]
    total_income = 0
    
    for income in incomes:
        income_data.append([
            income.date.strftime('%d-%b'),
            income.perticulers[:30],
            f"AED {income.amount:.2f}",
            income.bill_number
        ])
        total_income += income.amount
    
    income_data.append(['', 'Total:', f"AED {total_income:.2f}", ''])
    
    income_table = Table(income_data, colWidths=[1.2*inch, 3.5*inch, 1.5*inch, 1.8*inch])
    income_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#43A574')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6E6E6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 8)
    ]))
    
    elements.append(income_table)
    elements.append(Spacer(1, 20))
    
    # Expenses
    elements.append(Paragraph("EXPENSES", heading_style))
    
    expenses = Expense.objects.filter(date__gte=start, date__lte=end).order_by('date')
    expense_data = [['Date', 'Particulars', 'Amount', 'Bill Number']]
    total_expense = 0
    
    for expense in expenses:
        expense_data.append([
            expense.date.strftime('%d-%b'),
            expense.perticulers[:30],
            f"AED {expense.amount:.2f}",
            expense.bill_number
        ])
        total_expense += expense.amount
    
    expense_data.append(['', 'Total:', f"AED {total_expense:.2f}", ''])
    
    expense_table = Table(expense_data, colWidths=[1.2*inch, 3.5*inch, 1.5*inch, 1.8*inch])
    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D54395')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6E6E6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 8)
    ]))
    
    elements.append(expense_table)
    elements.append(Spacer(1, 20))
    
    # Summary
    elements.append(Paragraph("PERIOD SUMMARY", heading_style))
    
    total_receipts = float(total_fees) + total_income
    net_balance = total_receipts - total_expense
    
    summary_data = [
        ['Description', 'Amount'],
        ['Total Fee Collection', f"AED {total_fees:.2f}"],
        ['Total Other Income', f"AED {total_income:.2f}"],
        ['Total Receipts', f"AED {total_receipts:.2f}"],
        ['Total Expenses', f"AED {total_expense:.2f}"],
        ['Net Balance', f"AED {net_balance:.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFB804')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.yellow),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Financial_Report_{start}_to_{end}.pdf"'
    return response



def generate_fee_tracking_excel(fee_category_id, start_date, end_date):
    """Generate fee tracking Excel report from Payment and PaymentItem models"""
    try:
        fee_category = FeeCategory.objects.get(id=fee_category_id)
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{fee_category.name} Tracking"
        
        # School Header
        ws.merge_cells('A1:H1')
        ws['A1'] = "Blossom British School"
        ws['A1'].font = Font(size=16, bold=True, color="214888")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:H2')
        ws['A2'] = "Villa No 2 University Street, Ajman UAE"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:H3')
        ws['A3'] = f"Fee Tracking Report: {fee_category.name}"
        ws['A3'].font = Font(size=14, bold=True)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A4:H4')
        ws['A4'] = f"Period: {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"
        ws['A4'].alignment = Alignment(horizontal='center')
        
        row = 6
        
        # Headers
        headers = ['Student ID', 'Student Name', 'Class', 'Total Due', 'Paid Amount', 'Balance', 'Payment Date', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="214888", end_color="214888", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Get all payment items for this fee category within date range
        payment_items = PaymentItem.objects.filter(
            fee_category=fee_category,
            payment__payment_date__gte=start,
            payment__payment_date__lte=end,
            payment__payment_status='completed'
        ).select_related('payment', 'payment__student', 'payment__student__class_room')
        
        print(f"DEBUG: Found {payment_items.count()} payment items for fee category {fee_category.name}")
        
        # Group by student
        student_data = {}
        for item in payment_items:
            student = item.payment.student
            student_id = student.id
            
            if student_id not in student_data:
                student_data[student_id] = {
                    'student_id': student.student_id,
                    'student_name': student.get_full_name(),
                    'class_name': student.class_room.class_name if student.class_room else 'N/A',
                    'total_due': Decimal('0'),
                    'total_paid': Decimal('0'),
                    'last_payment_date': None,
                }
            
            # Add to totals
            student_data[student_id]['total_due'] += item.amount
            student_data[student_id]['total_paid'] += item.net_amount
            
            # Track latest payment date
            if item.payment.payment_date:
                if student_data[student_id]['last_payment_date'] is None:
                    student_data[student_id]['last_payment_date'] = item.payment.payment_date
                else:
                    if item.payment.payment_date > student_data[student_id]['last_payment_date']:
                        student_data[student_id]['last_payment_date'] = item.payment.payment_date
        
        print(f"DEBUG: Grouped data for {len(student_data)} students")
        
        total_due = Decimal('0')
        total_paid = Decimal('0')
        
        # Write data rows
        for student_id, data in sorted(student_data.items(), key=lambda x: x[1]['student_name']):
            balance = data['total_due'] - data['total_paid']
            payment_date = data['last_payment_date'].strftime('%d %b %Y') if data['last_payment_date'] else 'N/A'
            status = 'Paid' if balance <= 0 else 'Pending' if data['total_paid'] == 0 else 'Partial'
            
            ws.cell(row=row, column=1, value=data['student_id'])
            ws.cell(row=row, column=2, value=data['student_name'])
            ws.cell(row=row, column=3, value=data['class_name'])
            ws.cell(row=row, column=4, value=float(data['total_due']))
            ws.cell(row=row, column=5, value=float(data['total_paid']))
            ws.cell(row=row, column=6, value=float(balance))
            ws.cell(row=row, column=7, value=payment_date)
            ws.cell(row=row, column=8, value=status)
            
            # Color code status
            status_cell = ws.cell(row=row, column=8)
            if status == 'Paid':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status == 'Pending':
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            total_due += data['total_due']
            total_paid += data['total_paid']
            row += 1
        
        # Totals row
        row += 1
        ws.cell(row=row, column=3, value="TOTALS:").font = Font(bold=True, size=12)
        ws.cell(row=row, column=4, value=float(total_due)).font = Font(bold=True, size=12)
        ws.cell(row=row, column=5, value=float(total_paid)).font = Font(bold=True, size=12)
        ws.cell(row=row, column=6, value=float(total_due - total_paid)).font = Font(bold=True, size=12)
        
        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 16
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Fee_Tracking_{fee_category.name}_{start}_to_{end}.xlsx"'
        wb.save(response)
        return response
    
    except Exception as e:
        print(f"ERROR in generate_fee_tracking_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


def generate_fee_tracking_pdf(fee_category_id, start_date, end_date):
    """Generate fee tracking PDF report from Payment and PaymentItem models"""
    try:
        fee_category = FeeCategory.objects.get(id=fee_category_id)
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#214888'),
            alignment=TA_CENTER,
            spaceAfter=10
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#214888'),
            spaceAfter=10
        )
        
        # Header
        elements.append(Paragraph("Blossom British School", title_style))
        elements.append(Paragraph("Villa No 2 University Street, Ajman UAE", styles['Normal']))
        elements.append(Paragraph(f"Fee Tracking Report: {fee_category.name}", heading_style))
        elements.append(Paragraph(f"Period: {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Get all payment items for this fee category within date range
        payment_items = PaymentItem.objects.filter(
            fee_category=fee_category,
            payment__payment_date__gte=start,
            payment__payment_date__lte=end,
            payment__payment_status='completed'
        ).select_related('payment', 'payment__student', 'payment__student__class_room')
        
        print(f"DEBUG: Found {payment_items.count()} payment items for fee category {fee_category.name}")
        
        # Group by student
        student_data = {}
        for item in payment_items:
            student = item.payment.student
            student_id = student.id
            
            if student_id not in student_data:
                student_data[student_id] = {
                    'student_id': student.student_id,
                    'student_name': student.get_full_name(),
                    'class_name': student.class_room.class_name if student.class_room else 'N/A',
                    'total_due': Decimal('0'),
                    'total_paid': Decimal('0'),
                    'last_payment_date': None,
                }
            
            # Add to totals
            student_data[student_id]['total_due'] += item.amount
            student_data[student_id]['total_paid'] += item.net_amount
            
            # Track latest payment date
            if item.payment.payment_date:
                if student_data[student_id]['last_payment_date'] is None:
                    student_data[student_id]['last_payment_date'] = item.payment.payment_date
                else:
                    if item.payment.payment_date > student_data[student_id]['last_payment_date']:
                        student_data[student_id]['last_payment_date'] = item.payment.payment_date
        
        print(f"DEBUG: Grouped data for {len(student_data)} students")
        
        # Create table data
        table_data = [['Student ID', 'Name', 'Class', 'Due', 'Paid', 'Balance', 'Status']]
        total_due = Decimal('0')
        total_paid = Decimal('0')
        
        # Write data rows
        for student_id, data in sorted(student_data.items(), key=lambda x: x[1]['student_name']):
            balance = data['total_due'] - data['total_paid']
            status = 'Paid' if balance <= 0 else 'Pending' if data['total_paid'] == 0 else 'Partial'
            
            table_data.append([
                data['student_id'],
                data['student_name'][:25],
                data['class_name'],
                f"AED {data['total_due']:.2f}",
                f"AED {data['total_paid']:.2f}",
                f"AED {balance:.2f}",
                status
            ])
            
            total_due += data['total_due']
            total_paid += data['total_paid']
        
        # Add totals row
        table_data.append([
            '', 'TOTALS', '',
            f"AED {total_due:.2f}",
            f"AED {total_paid:.2f}",
            f"AED {(total_due - total_paid):.2f}",
            ''
        ])
        
        # Create table
        col_widths = [1.2*inch, 2*inch, 1*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1*inch]
        table = Table(table_data, colWidths=col_widths)
        
        # Apply styles
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#214888')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFB804')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8)
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Fee_Tracking_{fee_category.name}_{start}_to_{end}.pdf"'
        return response
    
    except Exception as e:
        print(f"ERROR in generate_fee_tracking_pdf: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


def generate_fee_tracking_report(request):
    """Generate fee tracking report by category"""
    if request.method == 'POST':
        fee_category_id = request.POST.get('fee_category')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        report_format = request.POST.get('format', 'excel')
        
        print(f"DEBUG: fee_category_id={fee_category_id}, start_date={start_date}, end_date={end_date}, format={report_format}")
        
        try:
            if report_format == 'excel':
                return generate_fee_tracking_excel(fee_category_id, start_date, end_date)
            else:
                return generate_fee_tracking_pdf(fee_category_id, start_date, end_date)
        except Exception as e:
            print(f"ERROR in generate_fee_tracking_report: {str(e)}")
            messages.error(request, f"Error generating report: {str(e)}")
            return redirect('reports_dashboard')
    
    return redirect('reports_dashboard')
# def generate_fee_tracking_excel(fee_category_id, start_date, end_date):
#     """Generate fee tracking Excel report"""
#     fee_category = FeeCategory.objects.get(id=fee_category_id)
#     start = datetime.strptime(start_date, '%Y-%m-%d').date()
#     end = datetime.strptime(end_date, '%Y-%m-%d').date()
    
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = f"{fee_category.name} Tracking"
    
#     # School Header
#     ws.merge_cells('A1:H1')
#     ws['A1'] = "Blossom British School"
#     ws['A1'].font = Font(size=16, bold=True, color="214888")
#     ws['A1'].alignment = Alignment(horizontal='center')
    
#     ws.merge_cells('A2:H2')
#     ws['A2'] = "Villa No 2 University Street, Ajman UAE"
#     ws['A2'].alignment = Alignment(horizontal='center')
    
#     ws.merge_cells('A3:H3')
#     ws['A3'] = f"Fee Tracking Report: {fee_category.name}"
#     ws['A3'].font = Font(size=14, bold=True)
#     ws['A3'].alignment = Alignment(horizontal='center')
    
#     ws.merge_cells('A4:H4')
#     ws['A4'] = f"Period: {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"
#     ws['A4'].alignment = Alignment(horizontal='center')
    
#     row = 6
    
#     # Headers
#     headers = ['Student ID', 'Student Name', 'Class', 'Total Due', 'Paid Amount', 'Balance', 'Payment Date', 'Status']
#     for col, header in enumerate(headers, 1):
#         cell = ws.cell(row=row, column=col, value=header)
#         cell.font = Font(bold=True, color="FFFFFF")
#         cell.fill = PatternFill(start_color="214888", end_color="214888", fill_type="solid")
#         cell.alignment = Alignment(horizontal='center')
#     row += 1
    
#     # Get all students with this fee category
#     fee_assignments = StudentFeeAssignment.objects.filter(
#         fee_structure__fee_category=fee_category,
#         is_active=True
#     ).select_related('student', 'fee_structure')
    
#     total_due = Decimal('0')
#     total_paid = Decimal('0')
    
#     for assignment in fee_assignments:
#         student = assignment.student
#         due_amount = assignment.get_final_amount()
        
#         # Calculate paid amount for this fee category
#         paid = PaymentItem.objects.filter(
#             payment__student=student,
#             fee_category=fee_category,
#             payment__payment_date__gte=start,
#             payment__payment_date__lte=end,
#             payment__payment_status='completed'
#         ).aggregate(total=Sum('net_amount'))['total'] or Decimal('0')
        
#         balance = due_amount - paid
        
#         # Get last payment date
#         last_payment = Payment.objects.filter(
#             student=student,
#             payment_items__fee_category=fee_category,
#             payment_date__gte=start,
#             payment_date__lte=end
#         ).order_by('-payment_date').first()
        
#         payment_date = last_payment.payment_date.strftime('%d %b %Y') if last_payment else 'N/A'
#         status = 'Paid' if balance <= 0 else 'Pending' if paid == 0 else 'Partial'
        
#         ws.cell(row=row, column=1, value=student.student_id)
#         ws.cell(row=row, column=2, value=student.get_full_name())
#         ws.cell(row=row, column=3, value=student.class_room.class_name if student.class_room else 'N/A')
#         ws.cell(row=row, column=4, value=float(due_amount))
#         ws.cell(row=row, column=5, value=float(paid))
#         ws.cell(row=row, column=6, value=float(balance))
#         ws.cell(row=row, column=7, value=payment_date)
#         ws.cell(row=row, column=8, value=status)
        
#         # Color code status
#         status_cell = ws.cell(row=row, column=8)
#         if status == 'Paid':
#             status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
#         elif status == 'Pending':
#             status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#         else:
#             status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
#         total_due += due_amount
#         total_paid += paid
#         row += 1
    
#     # Totals
#     row += 1
#     ws.cell(row=row, column=3, value="TOTALS:").font = Font(bold=True, size=12)
#     ws.cell(row=row, column=4, value=float(total_due)).font = Font(bold=True, size=12)
#     ws.cell(row=row, column=5, value=float(total_paid)).font = Font(bold=True, size=12)
#     ws.cell(row=row, column=6, value=float(total_due - total_paid)).font = Font(bold=True, size=12)
    
#     # Adjust column widths
#     for col in range(1, 9):
#         ws.column_dimensions[get_column_letter(col)].width = 16
    
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename="Fee_Tracking_{fee_category.name}_{start}_to_{end}.xlsx"'
#     wb.save(response)
#     return response